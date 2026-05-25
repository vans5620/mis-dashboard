"""
Run this script weekly to regenerate data.json from the latest IPS file.
Reads BOTH the 'Repository' tab (DPMS + NDPMS schemes) and the 'Navigate' tab
(AIF — Navigate AIF scheme) and merges them into a single record set.
Usage: python3 update_data.py
"""
import openpyxl, json, datetime, re, glob
from pathlib import Path

# Resolve the IPS file robustly — the script may be run from either the workspace
# subfolder OR from a separately-mounted allocate-mis-dashboard folder.
_candidates = [
    Path(__file__).parent.parent / "03 Client Repository IPS v2.xlsx",
    *[Path(p) / "03 Client Repository IPS v2.xlsx"
      for p in glob.glob("/sessions/*/mnt/Prashant Sureja*s files - Allocate")],
]
IPS_FILE = next((p for p in _candidates if p.exists()), _candidates[0])
OUT_FILE = Path(__file__).parent / "data.json"

SCHEME_MAP = {
    'Ionic Allocate Portfolio Aggressive':  'Allocate Aggressive',
    'Ionic Allocate Portfolio-Aggressive':  'Allocate Aggressive',
    'Ionic Allocate Portfolio Moderate':    'Allocate Moderate',
    'Ionic Allocate Portfolio-Moderate':    'Allocate Moderate',
    'Ionic Allocate Portfolio Equity':      'Allocate Equity',
    'Ionic Allocate Portfolio-Equity':      'Allocate Equity',
    'Ionic Liquid Approach DPMS':           'Liquid DPMS',
    'Ionic Allocate Select Portfolio':      'Allocate Select',
    'Ionic Large Value Portfolio':          'Large Value Portfolio',
    'Ionic Co-pilot Strategy':              'Co-pilot',
    'Ionic Copilot Strategy':               'Co-pilot',
    'Ionic Summit Portfolio Aggressive':    'Summit Aggressive',
    'Ionic Summit Portfolio Equity':        'Summit Equity',
    'Ionic Navigate AIF':                   'Navigate AIF',
}
DPMS  = ['Allocate Aggressive','Allocate Moderate','Allocate Equity','Liquid DPMS','Summit Aggressive','Summit Equity']
NDPMS = ['Allocate Select','Large Value Portfolio','Co-pilot']
AIF   = ['Navigate AIF']

RM_ALIASES = {
    'Karan Chandok':   ['karan chandok','karan chandhok'],
    'Prateek Chhabra': ['prateek chhabra','prateek chabra'],
    'Mani Sawhney':    ['mani sawhney','mani sawheny'],
    'Ishan Mishra':    ['ishan mishra','ishan mishra '],
}
ALIAS = {a.lower().strip(): canon for canon, lst in RM_ALIASES.items() for a in lst}

def norm_rm(raw):
    if not raw: return 'Unknown'
    k = str(raw).strip().lower()
    return ALIAS.get(k, str(raw).strip())

def map_trans(t):
    if t is None: return None
    s = str(t).strip().lower()
    if s in ('new activation', 'initial inflow', 'new client'): return 'New Activation'
    if s in ('top-up', 'top - up', 'top up'):                   return 'Top-up'
    return None

def parse_amt(v):
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s.startswith('='):
            expr = s[1:].replace(',', '')
            if re.fullmatch(r'[\d\s+\-*/().eE]+', expr):
                try: return float(eval(expr, {"__builtins__": {}}, {}))
                except: return None
            return None
        try: return float(s.replace(',', ''))
        except: return None
    return None

def to_month(raw_date):
    if hasattr(raw_date, 'year'):
        return raw_date.strftime('%Y-%m')
    if isinstance(raw_date, str):
        for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d'):
            try: return datetime.datetime.strptime(raw_date.strip(), fmt).strftime('%Y-%m')
            except: pass
    return 'unknown'

wb    = openpyxl.load_workbook(IPS_FILE, read_only=True, data_only=True)
ws    = wb['Repository']           # renamed from 'Client Summary' on 2026-05-04
ws_rm = wb['RM Mapping']
ws_nav = wb['Navigate']            # AIF — Navigate AIF (all rows implicitly = Ionic Navigate AIF)

rm_lead = {}
for r in ws_rm.iter_rows(values_only=True):
    if r[0] and r[1]: rm_lead[norm_rm(r[0])] = str(r[1]).strip()

records = []

# ── Repository tab (DPMS + NDPMS) ────────────────────────────────────────────
for r in ws.iter_rows(values_only=True):
    if not r[0]: continue
    if r[3] not in SCHEME_MAP: continue
    trans = map_trans(r[5])
    if not trans: continue
    amt = parse_amt(r[6])
    if amt is None or amt <= 0: continue
    date = r[7]
    month = to_month(date)
    if month == 'unknown': continue
    div = r[16] if r[16] in ('HNI','UHNI') else 'Other'
    rm  = norm_rm(r[13])
    ml  = rm_lead.get(rm, str(r[15]).strip() if r[15] else 'Unknown')
    records.append({
        'month': month, 'scheme': SCHEME_MAP[r[3]],
        'trans_type': trans, 'amount': amt, 'rm': rm,
        'market_lead': ml, 'division': div,
        'client_code': str(r[1]) if r[1] else None
    })

# ── Navigate tab (AIF — all rows implicitly Ionic Navigate AIF) ──────────────
# Columns: A=SrNo  B=ClientName/Code  C=?  D=Scheme  E=TxnType  F=Amount
#          G=ActivationDate  H=OtherDate  I=RM  J=MarketLead(formula)  K=Division(formula)
nav_count = 0
for i, r in enumerate(ws_nav.iter_rows(values_only=True), start=1):
    if i == 1: continue              # header
    if not r[0]: continue            # no serial number → empty row
    trans = map_trans(r[4])          # col E
    if not trans: continue
    amt = parse_amt(r[5])            # col F
    if amt is None or amt <= 0: continue
    month = to_month(r[6])           # col G = Activation Date (user-confirmed)
    if month == 'unknown': continue
    rm  = norm_rm(r[8])              # col I
    ml  = r[9] if r[9] else rm_lead.get(rm, 'Unknown')   # col J (formula result)
    ml  = str(ml).strip() if ml else 'Unknown'
    div = r[10] if r[10] in ('HNI','UHNI') else 'Other'  # col K
    records.append({
        'month': month, 'scheme': 'Navigate AIF',
        'trans_type': trans, 'amount': amt, 'rm': rm,
        'market_lead': ml, 'division': div,
        'client_code': None       # client codes not tracked for AIF
    })
    nav_count += 1

out = {
    'records': records, 'dpms': DPMS, 'ndpms': NDPMS, 'aif': AIF,
    'all_schemes': DPMS + NDPMS + AIF,
    'last_updated': datetime.date.today().isoformat()
}
with open(OUT_FILE, 'w') as f: json.dump(out, f)
total = sum(r['amount'] for r in records)
aif_total = sum(r['amount'] for r in records if r['scheme']=='Navigate AIF')
print(f"data.json updated — {len(records)} records, ₹{total/1e7:.2f} Cr, date: {out['last_updated']}")
print(f"  · Navigate AIF: {nav_count} records, ₹{aif_total/1e7:.2f} Cr")
)
